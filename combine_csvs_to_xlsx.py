#!/usr/bin/env python3
"""
Combine EnGenius export CSVs into a single Excel workbook.
Usage: python3 combine_csvs_to_xlsx.py <folder_path>
Requires: openpyxl
"""
import sys, os, csv, glob
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

SKIP_COLS = {"organization", "hierarchy_view"}

ABBREV = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY"
}


def tab_name(filename):
    name = os.path.splitext(filename)[0]
    for state, ab in ABBREV.items():
        if name.startswith(state):
            name = name.replace(state, ab, 1); break
    return name.replace("Wireless APs", "APs")[:31]


def try_date(val):
    """Try to parse a date string and return a datetime object, or None."""
    if not val: return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try: return datetime.strptime(val.strip(), fmt)
        except: continue
    return None


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 combine_csvs_to_xlsx.py <csv_folder>"); sys.exit(1)

    folder = sys.argv[1]
    if not os.path.isdir(folder):
        print(f"ERROR: '{folder}' not a directory"); sys.exit(1)

    csv_files = sorted(glob.glob(os.path.join(folder, "*.csv")))
    if not csv_files:
        print(f"No CSVs in {folder}"); sys.exit(1)

    print(f"Found {len(csv_files)} CSV(s)")

    wb = Workbook()
    wb.remove(wb.active)

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    border = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    expired_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    expiring_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    online_font = Font(color="006100")
    offline_font = Font(color="9C0006")

    for csv_path in csv_files:
        fname = os.path.basename(csv_path)
        tname = tab_name(fname)
        print(f"  {fname} -> [{tname}]")
        ws = wb.create_sheet(title=tname)

        with open(csv_path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)

            # Figure out which column indices to keep
            keep = [i for i, h in enumerate(header) if h.strip().lower() not in SKIP_COLS]

            # Write header
            for out_col, src_col in enumerate(keep, 1):
                cell = ws.cell(row=1, column=out_col, value=header[src_col])
                cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

            # Track which output columns are date columns
            date_cols = set()
            for out_col, src_col in enumerate(keep, 1):
                if "expiration" in header[src_col].lower() and "days" not in header[src_col].lower():
                    date_cols.add(out_col)

            # Write data
            for row_idx, row in enumerate(reader, 2):
                for out_col, src_col in enumerate(keep, 1):
                    val = row[src_col] if src_col < len(row) else ""
                    cell = ws.cell(row=row_idx, column=out_col)
                    cell.border = border

                    # Date columns -> actual date with short format
                    if out_col in date_cols:
                        dt = try_date(val)
                        if dt:
                            cell.value = dt
                            cell.number_format = "M/D/YYYY"
                        else:
                            cell.value = val
                    else:
                        cell.value = val

                    # Color coding
                    if val == "EXPIRED": cell.fill = expired_fill
                    elif val == "EXPIRING SOON": cell.fill = expiring_fill
                    elif val == "Online": cell.font = online_font
                    elif val == "Offline": cell.font = offline_font

        # Auto-width
        for col in range(1, ws.max_column + 1):
            mx = max((len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[get_column_letter(col)].width = min(mx + 3, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    out_path = os.path.join(os.path.dirname(folder.rstrip("/\\")), f"EnGenius_Device_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(out_path)
    print(f"\nSaved: {out_path} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
